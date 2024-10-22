import requests
import csv
from openpyxl import load_workbook 
from tqdm import tqdm

api_url = 'http://localhost:11434/api/generate'
model_name = 'qwen2:72b'
data_name = 'text2sql.xlsx'
sheet_name = 'Sheet1'
output_name = 'newq.csv'

PROMPT = '''
你是一个数据库sql分析专家，通过提供的数据库表结构，分析用户的问题以及回答用户问题所需的sql语句。
数据库表结构如下：
{table_info}
用户问题如下：
{question}
回答用户的sql语句如下：
{sql}
你需要对用户的问题进行变化，并且以如下格式进行回答：
```question1 xxx```
```question2 xxx```
```question3 xxx```
```question4 xxx```
```question5 xxx```
```question6 xxx```
```question7 xxx```
```question8 xxx```
```question9 xxx```
```question10 xxx```
回答要求：
1，写出原question变换后的十个问题，question1-10后面直接写出变化后的问题，不要有其他分析的内容
2，变换后的问题可以是原问题的同义词替换，例如’一个月‘和’30天‘相互替换，’一周‘和’7天‘相互替换
3，变换后的问题也可以是原问题中名词的位置交换，例如‘绍兴市今年的事件量’与‘今年绍兴市的事件量’互换
4，变换后的问题与原问题的语意基本相同，原sql代码依旧可以用来回答变换后的问题
请按要求进行回答：
'''

workbook = load_workbook(filename=data_name)
sheet = workbook[sheet_name]

datas = []

for row in tqdm(sheet.iter_rows(min_row=2, values_only=True)):
    prompt = PROMPT.format(table_info=row[1], question=row[2], sql=row[3])
    print('prompt:', prompt)

    request_data = {
        'model': model_name,
        'prompt': prompt,
        'stream': False,
        'options': {
            'temprature': 0.8
        }
    }

    response = requests.post(api_url, json=request_data)
    output = response.json()['response']
    print('output:', output)

    o = [r for r in row]
    o.append(row[2])
    o.append(row[3])
    datas.append(o)

    for i in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        q2 = output.split(f'```question{i} ')[-1].split('```')[0]
        if q2 and q2[0] == '\n': q2 = q2[1:]
        if q2 and q2[-1] == '\n': q2 = q2[:-1]
    
        o = [r for r in row]
        o.append(q2)
        o.append(row[3])
        datas.append(o)

    print('------')

with open(output_name, 'w', encoding='UTF8') as f:
    writer = csv.writer(f)
    writer.writerows(datas)

print('done')
